from collections import Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import streamlit as st
from tempfile import NamedTemporaryFile
import zulip
from stqdm import stqdm
from time import strftime as stt
import os


def disc_apply(row):
    a = row['cat']
    b = row['disc']
    c = []
    for i, j in enumerate(a):
        for k in range(b[i]):
            c.append(j)
    return c


def proc_wes(data, warehouse, matches):
    try:
        df = data.copy()
        if warehouse:
            df = df[df['Warehouse Number'] == warehouse]

        store_dict = dict(zip(df['Store Number'], df['Store Name']))
        area_dict = dict(zip(df['Store Number'], df['Area Name']))
        # cat_dict = dict(zip(df['Product No'], df['Product Description']))

        df['from-doc-to'] = df['Warehouse Number'].astype(str) + '-' + df['Delivery Number'].astype(str) + '-' + df['Store Number'].astype(str)
        df['disc'] = df['Voiced Quantity'] - df['Advised Quantity']
        df.rename(columns={'Product No': 'cat', 'Product Description': 'desc', 'Short / Over': 's-o'}, inplace=True)
        df = df[['from-doc-to', 'cat', 'disc']]

        ss = df.copy()
        ss = ss[ss['disc'] < 0]
        ss['disc'] = ss['disc'].abs()
        ss2 = ss.groupby(['from-doc-to'], as_index=False).agg(lambda x: x.tolist())
        ss2['lines'] = ss2.apply(disc_apply, axis=1)
        del ss2['cat']
        del ss2['disc']

        oo = df.copy()
        oo = oo[df['disc'] > 0]
        oo2 = oo.groupby(['from-doc-to'], as_index=False).agg(lambda x: x.tolist())
        oo2['lines'] = oo2.apply(disc_apply, axis=1)
        del oo2['cat']
        del oo2['disc']

        shorts = ss2['lines'].tolist()
        overs = oo2['lines'].tolist()

        shorts = [x if len(x) >= matches else [] for x in shorts]
        overs = [x if len(x) >= matches else [] for x in overs]
    except Exception as e:
        print('proc_wes 1 -', e)

    try:
        ###########################################################################
        #   Statistics - Shorts
        ###########################################################################
        dfs = ss2.copy()
        dfs['dc'] = dfs['from-doc-to'].str[:3]
        dfs1 = dfs.copy()
        dfs5 = dfs.copy()
        dfs = dfs.groupby(['dc'], as_index=False).agg('count')
        del dfs['from-doc-to']
        dfs.rename(columns={'lines': 'all shorts'}, inplace=True)

        dfs1['disc_count'] = dfs1['lines'].str.len()
        dfs1 = dfs1[dfs1['disc_count'] == 1]
        dfs1 = dfs1.groupby(['dc'], as_index=False).agg('count')
        del dfs1['from-doc-to']
        del dfs1['lines']
        dfs1.rename(columns={'disc_count': 'single shorts'}, inplace=True)

        dfs5['disc_count'] = dfs5['lines'].str.len()
        dfs5 = dfs5[dfs5['disc_count'] >= 5]
        dfs5 = dfs5.groupby(['dc'], as_index=False).agg('count')
        del dfs5['from-doc-to']
        del dfs5['lines']
        dfs5.rename(columns={'disc_count': 'shorts >=5'}, inplace=True)

        del dfs1['dc']
        del dfs5['dc']
        df_shorts = pd.concat([dfs, dfs1, dfs5], axis=1)

        ###########################################################################
        #   Statistics - Overs
        ###########################################################################
        dfo = oo2.copy()
        dfo['dc'] = dfo['from-doc-to'].str[:3]
        dfo1 = dfo.copy()
        dfo5 = dfo.copy()
        dfo = dfo.groupby(['dc'], as_index=False).agg('count')
        del dfo['from-doc-to']
        dfo.rename(columns={'lines': 'all overs'}, inplace=True)

        dfo1['disc_count'] = dfo1['lines'].str.len()
        dfo1 = dfo1[dfo1['disc_count'] == 1]
        dfo1 = dfo1.groupby(['dc'], as_index=False).agg('count')
        del dfo1['from-doc-to']
        del dfo1['lines']
        dfo1.rename(columns={'disc_count': 'single overs'}, inplace=True)

        dfo5['disc_count'] = dfo5['lines'].str.len()
        dfo5 = dfo5[dfo5['disc_count'] >= 5]
        dfo5 = dfo5.groupby(['dc'], as_index=False).agg('count')
        del dfo5['from-doc-to']
        del dfo5['lines']
        dfo5.rename(columns={'disc_count': 'overs >=5'}, inplace=True)

        del dfo1['dc']
        del dfo5['dc']
        df_overs = pd.concat([dfo, dfo1, dfo5], axis=1)

        del df_overs['dc']
        df_stats = pd.concat([df_shorts, df_overs], axis=1)
        df_stats = df_stats[['dc', 'all shorts', 'all overs', 'single shorts', 'single overs', 'shorts >=5', 'overs >=5']]
        df_stats.columns = ['DC', 'All Shorts', 'All Overs', 'Single Qty Shorts', 'Single Qty Overs', 'Shorts Qty>=5', 'Overs Qty>=5']
    except Exception as e:
        print('proc_wes 2 -', e)

    try:
        ###########################################################################
        #   Matches
        ###########################################################################
        result = []
        for i, s in enumerate(stqdm(shorts, ncols=100)):
            for j, o in enumerate(overs):
                magic = list((Counter(o) & Counter(s)).elements())
                if len(magic) >= matches:
                    x = ss2['from-doc-to'].iloc[i].split('-')
                    y = oo2['from-doc-to'].iloc[j].split('-')

                    result.append([x[0], x[2], store_dict.get(int(x[2])), x[1], len(shorts[i]),
                                   y[0], y[2], store_dict.get(int(y[2])), y[1], len(overs[j]),
                                   x[2] == y[2],
                                   area_dict.get(int(x[2])) == area_dict.get(int(y[2])),
                                   x[0] == y[0],
                                   len(magic),
                                   # f"{len(magic)/len(shorts[i])*100}%",
                                   f"{len(magic)/len(shorts[i])}",
                                   # f"{len(magic)/len(overs[j])*100}%"])
                                   f"{len(magic)/len(overs[j])}"])

                    columns = ['DC', 'Store', 'Name', 'Delivery', 'Short Qty',
                               'DC_', 'Store_', 'Name_', 'Delivery_', 'Over Qty',
                               'Same Store',
                               'Same Area',
                               'Same DC',
                               'Matching Qty',
                               '% vs Short',
                               '% vs Over']
    except Exception as e:
        print('proc_wes 3 -', e)

    try:
        xxx = pd.DataFrame(result, columns=columns).sort_values(by=['Matching Qty'], ascending=False)
        xxx_dict = {'DC': int,
                    'Store': int,
                    'Delivery': int,
                    'DC_': int,
                    'Store_': int,
                    'Delivery_': int,
                    '% vs Short': float,
                    '% vs Over': float
                    }
        xxx = xxx.astype(xxx_dict)
        df_stats = df_stats.astype({'DC': int})
        return xxx, df_stats
    except Exception as e:
        print('proc_wes 4 -', e)


def to_xlsx(matches_table, stats_table):
    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Statistics'
        ws2 = wb.create_sheet(title='Matches')

        for r in dataframe_to_rows(stats_table, index=False, header=True):
            ws1.append(r)

        for r in dataframe_to_rows(matches_table, index=False, header=True):
            ws2.append(r)

        table = Table(displayName="Table1", ref="A1:" + get_column_letter(ws1.max_column) + str(ws1.max_row))
        table2 = Table(displayName="Table2", ref="A1:" + get_column_letter(ws2.max_column) + str(ws2.max_row))

        style = TableStyleInfo(name="TableStyleMedium5", showFirstColumn=True, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        style2 = TableStyleInfo(name="TableStyleMedium5", showFirstColumn=False, showLastColumn=False,
                                showRowStripes=False, showColumnStripes=False)
        table2.tableStyleInfo = style2

        ws1.column_dimensions["A"].width = 9
        ws1.column_dimensions["B"].width = 18
        ws1.column_dimensions["C"].width = 18
        ws1.column_dimensions["D"].width = 18
        ws1.column_dimensions["E"].width = 18
        ws1.column_dimensions["F"].width = 18
        ws1.column_dimensions["G"].width = 18

        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 9
        ws2.column_dimensions["C"].width = 23
        ws2.column_dimensions["D"].width = 11
        ws2.column_dimensions["E"].width = 12
        ws2.column_dimensions["F"].width = 8
        ws2.column_dimensions["G"].width = 9
        ws2.column_dimensions["H"].width = 23
        ws2.column_dimensions["I"].width = 11
        ws2.column_dimensions["J"].width = 12

        ws2.column_dimensions["K"].width = 13
        ws2.column_dimensions["L"].width = 12.5
        ws2.column_dimensions["M"].width = 11

        ws2.column_dimensions["N"].width = 15
        ws2.column_dimensions["O"].width = 12
        ws2.column_dimensions["P"].width = 12

        for row in ws2[f"A2:D{ws2.max_row}"]:
            for cell in row:
                cell.style = '20 % - Accent4'
        for row in ws2[f"E2:E{ws2.max_row}"]:
            for cell in row:
                cell.style = '40 % - Accent4'
        for row in ws2[f"F2:I{ws2.max_row}"]:
            for cell in row:
                cell.style = '20 % - Accent4'
        for row in ws2[f"J2:J{ws2.max_row}"]:
            for cell in row:
                cell.style = '40 % - Accent4'
        for row in ws2[f"K2:M{ws2.max_row}"]:
            for cell in row:
                cell.style = '20 % - Accent4'
        for row in ws2[f"N2:N{ws2.max_row}"]:
            for cell in row:
                cell.style = '40 % - Accent4'
        for row in ws2[f"O2:P{ws2.max_row}"]:
            for cell in row:
                cell.style = '20 % - Accent4'

        for row in ws1[2:ws1.max_row]:
            cella = row[0]
            cella.alignment = Alignment(horizontal='left')

        for row in ws2[2:ws2.max_row]:
            cella = row[0]
            cella.alignment = Alignment(horizontal='left')
            cellb = row[1]
            cellb.alignment = Alignment(horizontal='left')
            celld = row[3]
            celld.alignment = Alignment(horizontal='left')
            celle = row[4]
            celle.alignment = Alignment(horizontal='center')
            cellf = row[5]
            cellf.alignment = Alignment(horizontal='left')
            cellg = row[6]
            cellg.alignment = Alignment(horizontal='left')
            celli = row[8]
            celli.alignment = Alignment(horizontal='left')
            cellj = row[9]
            cellj.alignment = Alignment(horizontal='center')
            celln = row[13]
            celln.alignment = Alignment(horizontal='center')

            cell1 = row[14]
            cell1.number_format = '0.00%'
            cell1.alignment = Alignment(horizontal='left')
            cell2 = row[15]
            cell2.number_format = '0.00%'
            cell2.alignment = Alignment(horizontal='left')

        ws1.add_table(table)
        ws2.add_table(table2)

        c = ws2['A2']
        ws2.freeze_panes = c

        return wb
    except Exception as e:
        print('to_xlsx', e)


def zulip_msg():
    try:
        zulip.Client(api_key=os.environ.get('msg_key'),
                     email=os.environ.get('msg_mail'),
                     site=os.environ.get('msg_site')).send_message(
            {"type": "private", "to": [int(os.environ.get('msg_to'))],
             "content": f"WES Report ran at {stt('%H:%M:%S on %d-%m-%y')}"})
    except Exception as e:
        print('zulip_msg', e)


@st.cache_data
def action(data_from_ul):
    try:
        print('lilaliba')
        a, b = proc_wes(data=pd.read_excel(data_from_ul), warehouse=None, matches=min_match)
        wb_obj = to_xlsx(a, b)

        with NamedTemporaryFile() as tmp:
            tmp.close()
            wb_obj.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                f.seek(0)
                new_file_object = f.read()
                zulip_msg()
                return new_file_object
    except Exception as e:
        print('action', e)


st.set_page_config(
    page_title='WES Matching',
    page_icon='🔗',
)

st.title("WES Short-Over Matching Microapp")
st.subheader('Created by Daniel Matyasi')
st.write('')
st.write('')
st.write('')
st.write('')
min_match = st.number_input(label='Set minimum matches', value=5, min_value=1, max_value=25)
st.write('')
st.write('')
st.write('')
st.write('')
uploaded_file = st.file_uploader(label='Remember to remove the empty top row before uploading WES data file!')

if uploaded_file is not None:
    to_dl = action(uploaded_file)

    st.download_button(
        label="Download Report",
        data=to_dl,
        file_name='WES Stats and Matches' + '.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
